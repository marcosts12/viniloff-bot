import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

ARQUIVO_HISTORICO = "/home/claude/historico_viniloff.json"
ARQUIVO_SAIDA = "/home/claude/viniloff_produtos.xlsx"

# Cores VinilOFF
COR_FUNDO_HEADER = "2C1A0E"
COR_TEXTO_HEADER = "E8DDD0"
COR_ACCENT       = "C8541A"
COR_VERDE        = "1A3A2A"
COR_LINHA_PAR    = "F5F0EB"
COR_PROMO        = "FFF3EC"
COR_PROMO_FORTE  = "FFE0CC"

def borda_fina():
    lado = Side(style="thin", color="DDCCBB")
    return Border(left=lado, right=lado, top=lado, bottom=lado)

def gerar_planilha():
    if not os.path.exists(ARQUIVO_HISTORICO):
        print("Arquivo historico_viniloff.json não encontrado!")
        return

    with open(ARQUIVO_HISTORICO, "r", encoding="utf-8") as f:
        historico = json.load(f)

    wb = Workbook()

    # ── ABA 1: TODOS OS PRODUTOS ──
    ws1 = wb.active
    ws1.title = "Todos os Produtos"

    # Título principal
    ws1.merge_cells("A1:H1")
    ws1["A1"] = "VinilOFF — Produtos Monitorados"
    ws1["A1"].font = Font(name="Arial", bold=True, size=14, color=COR_TEXTO_HEADER)
    ws1["A1"].fill = PatternFill("solid", fgColor=COR_FUNDO_HEADER)
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 30

    # Subtítulo
    ws1.merge_cells("A2:H2")
    ws1["A2"] = f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')} · {len(historico)} produtos"
    ws1["A2"].font = Font(name="Arial", size=10, color="7A5C4A")
    ws1["A2"].fill = PatternFill("solid", fgColor="3D2510")
    ws1["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[2].height = 20

    # Cabeçalhos
    headers = ["#", "Produto", "Preço Atual", "Preço Máximo", "Preço Mínimo", "Desconto %", "Status", "Link Amazon"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=3, column=col, value=h)
        cell.font = Font(name="Arial", bold=True, size=10, color=COR_TEXTO_HEADER)
        cell.fill = PatternFill("solid", fgColor=COR_ACCENT)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = borda_fina()
    ws1.row_dimensions[3].height = 22

    # Ordena por maior desconto
    itens = []
    for asin, d in historico.items():
        preco_atual  = d.get("preco_atual", 0)
        preco_maximo = d.get("preco_maximo", preco_atual)
        preco_minimo = d.get("preco_minimo", preco_atual)
        desconto     = round(((preco_maximo - preco_atual) / preco_maximo) * 100) if preco_maximo > 0 else 0
        itens.append({
            "asin": asin,
            "nome": d.get("nome", ""),
            "preco_atual": preco_atual,
            "preco_maximo": preco_maximo,
            "preco_minimo": preco_minimo,
            "desconto": desconto,
            "url": d.get("url", ""),
            "atualizado": d.get("atualizado", ""),
        })
    itens.sort(key=lambda x: x["desconto"], reverse=True)

    # Dados
    for i, item in enumerate(itens):
        row = i + 4
        eh_par = i % 2 == 0

        # Cor da linha
        if item["desconto"] >= 20:
            cor_linha = COR_PROMO_FORTE
        elif item["desconto"] >= 10:
            cor_linha = COR_PROMO
        else:
            cor_linha = COR_LINHA_PAR if eh_par else "FFFFFF"

        fill = PatternFill("solid", fgColor=cor_linha)

        # Número
        c = ws1.cell(row=row, column=1, value=i+1)
        c.font = Font(name="Arial", size=9, color="7A5C4A")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = fill
        c.border = borda_fina()

        # Nome
        c = ws1.cell(row=row, column=2, value=item["nome"])
        c.font = Font(name="Arial", size=10, bold=(item["desconto"] >= 10))
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.fill = fill
        c.border = borda_fina()

        # Preço atual
        c = ws1.cell(row=row, column=3, value=item["preco_atual"])
        c.font = Font(name="Arial", size=10, bold=True, color=COR_ACCENT if item["desconto"] >= 10 else "000000")
        c.number_format = 'R$ #,##0.00'
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = fill
        c.border = borda_fina()

        # Preço máximo
        c = ws1.cell(row=row, column=4, value=item["preco_maximo"])
        c.font = Font(name="Arial", size=10, color="999999", strike=(item["desconto"] >= 10))
        c.number_format = 'R$ #,##0.00'
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = fill
        c.border = borda_fina()

        # Preço mínimo
        c = ws1.cell(row=row, column=5, value=item["preco_minimo"])
        c.font = Font(name="Arial", size=10, color="1A3A2A")
        c.number_format = 'R$ #,##0.00'
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = fill
        c.border = borda_fina()

        # Desconto %
        desconto_formula = f"=(D{row}-C{row})/D{row}"
        c = ws1.cell(row=row, column=6, value=desconto_formula)
        c.font = Font(name="Arial", size=10, bold=(item["desconto"] >= 10),
                     color="C8541A" if item["desconto"] >= 10 else "666666")
        c.number_format = '0%'
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = fill
        c.border = borda_fina()

        # Status
        if item["desconto"] >= 20:
            status = "🔥 PROMOÇÃO"
        elif item["desconto"] >= 10:
            status = "📉 Em queda"
        elif item["desconto"] > 0:
            status = "↓ Leve queda"
        else:
            status = "💤 Normal"
        c = ws1.cell(row=row, column=7, value=status)
        c.font = Font(name="Arial", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = fill
        c.border = borda_fina()

        # Link
        c = ws1.cell(row=row, column=8, value=item["url"])
        c.font = Font(name="Arial", size=9, color="1155CC", underline="single")
        c.hyperlink = item["url"]
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = fill
        c.border = borda_fina()

        ws1.row_dimensions[row].height = 20

    # Larguras das colunas
    ws1.column_dimensions["A"].width = 5
    ws1.column_dimensions["B"].width = 40
    ws1.column_dimensions["C"].width = 14
    ws1.column_dimensions["D"].width = 14
    ws1.column_dimensions["E"].width = 14
    ws1.column_dimensions["F"].width = 12
    ws1.column_dimensions["G"].width = 16
    ws1.column_dimensions["H"].width = 35

    # Congela cabeçalho
    ws1.freeze_panes = "A4"

    # ── ABA 2: APENAS PROMOÇÕES ──
    ws2 = wb.create_sheet("🔥 Promoções Ativas")

    ws2.merge_cells("A1:H1")
    ws2["A1"] = "VinilOFF — Promoções Ativas (desconto ≥ 10%)"
    ws2["A1"].font = Font(name="Arial", bold=True, size=14, color=COR_TEXTO_HEADER)
    ws2["A1"].fill = PatternFill("solid", fgColor=COR_ACCENT)
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 30

    headers2 = ["#", "Produto", "Preço Atual", "Preço Máximo", "Desconto", "Economia", "Status", "Link Amazon"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=2, column=col, value=h)
        cell.font = Font(name="Arial", bold=True, size=10, color=COR_TEXTO_HEADER)
        cell.fill = PatternFill("solid", fgColor=COR_FUNDO_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = borda_fina()
    ws2.row_dimensions[2].height = 22

    promocoes = [x for x in itens if x["desconto"] >= 10]
    for i, item in enumerate(promocoes):
        row = i + 3
        fill = PatternFill("solid", fgColor=COR_PROMO_FORTE if item["desconto"] >= 20 else COR_PROMO)

        ws2.cell(row=row, column=1, value=i+1).font = Font(name="Arial", size=9)
        ws2.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws2.cell(row=row, column=1).fill = fill
        ws2.cell(row=row, column=1).border = borda_fina()

        c = ws2.cell(row=row, column=2, value=item["nome"])
        c.font = Font(name="Arial", size=10, bold=True)
        c.fill = fill; c.border = borda_fina()
        c.alignment = Alignment(vertical="center", wrap_text=True)

        c = ws2.cell(row=row, column=3, value=item["preco_atual"])
        c.font = Font(name="Arial", size=11, bold=True, color=COR_ACCENT)
        c.number_format = 'R$ #,##0.00'
        c.fill = fill; c.border = borda_fina()
        c.alignment = Alignment(horizontal="center", vertical="center")

        c = ws2.cell(row=row, column=4, value=item["preco_maximo"])
        c.font = Font(name="Arial", size=10, color="999999", strike=True)
        c.number_format = 'R$ #,##0.00'
        c.fill = fill; c.border = borda_fina()
        c.alignment = Alignment(horizontal="center", vertical="center")

        c = ws2.cell(row=row, column=5, value=f"=(D{row}-C{row})/D{row}")
        c.font = Font(name="Arial", size=10, bold=True, color=COR_ACCENT)
        c.number_format = '0%'
        c.fill = fill; c.border = borda_fina()
        c.alignment = Alignment(horizontal="center", vertical="center")

        c = ws2.cell(row=row, column=6, value=f"=D{row}-C{row}")
        c.font = Font(name="Arial", size=10, bold=True, color="1A3A2A")
        c.number_format = 'R$ #,##0.00'
        c.fill = fill; c.border = borda_fina()
        c.alignment = Alignment(horizontal="center", vertical="center")

        status = "🔥 PROMOÇÃO FORTE" if item["desconto"] >= 20 else "📉 Em promoção"
        c = ws2.cell(row=row, column=7, value=status)
        c.font = Font(name="Arial", size=10, bold=True)
        c.fill = fill; c.border = borda_fina()
        c.alignment = Alignment(horizontal="center", vertical="center")

        c = ws2.cell(row=row, column=8, value=item["url"])
        c.font = Font(name="Arial", size=9, color="1155CC", underline="single")
        c.hyperlink = item["url"]
        c.fill = fill; c.border = borda_fina()
        c.alignment = Alignment(horizontal="center", vertical="center")

        ws2.row_dimensions[row].height = 20

    ws2.column_dimensions["A"].width = 5
    ws2.column_dimensions["B"].width = 42
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 14
    ws2.column_dimensions["E"].width = 12
    ws2.column_dimensions["F"].width = 14
    ws2.column_dimensions["G"].width = 20
    ws2.column_dimensions["H"].width = 35
    ws2.freeze_panes = "A3"

    # ── ABA 3: RESUMO ──
    ws3 = wb.create_sheet("📊 Resumo")

    ws3.merge_cells("A1:D1")
    ws3["A1"] = "VinilOFF — Resumo do Monitoramento"
    ws3["A1"].font = Font(name="Arial", bold=True, size=14, color=COR_TEXTO_HEADER)
    ws3["A1"].fill = PatternFill("solid", fgColor=COR_FUNDO_HEADER)
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 30

    total = len(itens)
    em_promo = len([x for x in itens if x["desconto"] >= 10])
    promo_forte = len([x for x in itens if x["desconto"] >= 20])
    normal = total - em_promo

    resumo = [
        ("Total de produtos monitorados", total),
        ("Em promoção (≥ 10% de desconto)", em_promo),
        ("Promoção forte (≥ 20% de desconto)", promo_forte),
        ("Sem promoção no momento", normal),
        ("Gerado em", datetime.now().strftime("%d/%m/%Y %H:%M")),
    ]

    for i, (label, valor) in enumerate(resumo):
        row = i + 3
        c = ws3.cell(row=row, column=1, value=label)
        c.font = Font(name="Arial", size=11)
        c.fill = PatternFill("solid", fgColor=COR_LINHA_PAR if i % 2 == 0 else "FFFFFF")
        c.alignment = Alignment(vertical="center")
        c.border = borda_fina()

        c = ws3.cell(row=row, column=2, value=valor)
        c.font = Font(name="Arial", size=11, bold=True, color=COR_ACCENT)
        c.fill = PatternFill("solid", fgColor=COR_LINHA_PAR if i % 2 == 0 else "FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = borda_fina()
        ws3.row_dimensions[row].height = 22

    ws3.column_dimensions["A"].width = 38
    ws3.column_dimensions["B"].width = 20

    wb.save(ARQUIVO_SAIDA)
    print(f"Planilha salva: {ARQUIVO_SAIDA}")
    print(f"Total: {total} produtos | Em promoção: {em_promo} | Promoção forte: {promo_forte}")

gerar_planilha()
