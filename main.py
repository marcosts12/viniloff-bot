"""
VinilOFF — Bot Anti-Bloqueio com Telegram
==========================================
Versão melhorada com proteções contra bloqueio da Amazon.

COMO USAR:
1. pip install requests beautifulsoup4
2. Configure TELEGRAM_TOKEN e TELEGRAM_CHAT_ID abaixo
3. python viniloff_bot_auto.py
"""

import subprocess
import sys
subprocess.check_call([sys.executable, "-m", "pip", "install", "requests", "beautifulsoup4", "openpyxl"],
                      stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)

import requests
from bs4 import BeautifulSoup
import json
import time
import os
import random
import threading
from datetime import datetime

# ─────────────────────────────────────────────
#  CONFIGURAÇÕES
# ─────────────────────────────────────────────

TELEGRAM_TOKEN    = "8944782842:AAHMfyHMKSijzbby1lc-hNplMMGPu7BnP4s"
TELEGRAM_CHAT_ID  = 7484525336

DESCONTO_MINIMO   = 10    # % mínimo de queda para alertar
INTERVALO_MINUTOS = 20    # minutos entre varreduras
PAGINAS_VARRER    = 5     # páginas por categoria
ARQUIVO_HISTORICO = "historico_viniloff.json"

# ─────────────────────────────────────────────
#  CATEGORIAS
# ─────────────────────────────────────────────

CATEGORIAS = [
    {
        "nome": "Vinil — Todos",
        "url": "https://www.amazon.com.br/s?i=music&rh=n%3A19549018011&s=price-asc-rank"
    },
    {
        "nome": "Vinil — Pop",
        "url": "https://www.amazon.com.br/s?i=music&rh=n%3A19549018011%2Cn%3A6309293011&s=price-asc-rank"
    },
    {
        "nome": "Vinil — Rock",
        "url": "https://www.amazon.com.br/s?i=music&rh=n%3A19549018011%2Cn%3A6309285011&s=price-asc-rank"
    },
    {
        "nome": "Vinil — MPB",
        "url": "https://www.amazon.com.br/s?i=music&rh=n%3A19549018011%2Cn%3A6309271011&s=price-asc-rank"
    },
]

# ─────────────────────────────────────────────
#  PROTEÇÕES ANTI-BLOQUEIO
# ─────────────────────────────────────────────

# Lista grande de user agents reais
USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:121.0) Gecko/20100101 Firefox/121.0",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:120.0) Gecko/20100101 Firefox/120.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.2 Safari/605.1.15",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.1 Safari/605.1.15",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:121.0) Gecko/20100101 Firefox/121.0",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36 Edg/120.0.0.0",
]

# Contador de bloqueios consecutivos
bloqueios_consecutivos = 0
MAX_BLOQUEIOS = 3  # após 3 bloqueios seguidos avisa no Telegram


def log(msg):
    print(f"[{datetime.now().strftime('%d/%m %H:%M:%S')}] {msg}", flush=True)


def headers_aleatorios():
    """Gera headers realistas e variados a cada requisição"""
    accept_languages = [
        "pt-BR,pt;q=0.9,en-US;q=0.8,en;q=0.7",
        "pt-BR,pt;q=0.9,en;q=0.8",
        "pt-BR,pt;q=0.8,en-US;q=0.5,en;q=0.3",
    ]
    return {
        "User-Agent": random.choice(USER_AGENTS),
        "Accept-Language": random.choice(accept_languages),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
        "Accept-Encoding": "gzip, deflate, br",
        "DNT": random.choice(["1", "0"]),
        "Connection": "keep-alive",
        "Upgrade-Insecure-Requests": "1",
        "Sec-Fetch-Dest": "document",
        "Sec-Fetch-Mode": "navigate",
        "Sec-Fetch-Site": random.choice(["none", "same-origin"]),
        "Cache-Control": random.choice(["max-age=0", "no-cache"]),
    }


def pausa_humana(minimo=3, maximo=9):
    """Pausa aleatória que imita comportamento humano"""
    pausa = random.uniform(minimo, maximo)
    # Às vezes faz uma pausa mais longa como se estivesse lendo
    if random.random() < 0.1:
        pausa += random.uniform(5, 15)
    time.sleep(pausa)


def enviar_telegram(msg):
    try:
        url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
        r = requests.post(url, json={
            "chat_id": TELEGRAM_CHAT_ID,
            "text": msg,
            "parse_mode": "HTML",
            "disable_web_page_preview": False,
        }, timeout=10)
        return r.status_code == 200
    except Exception as e:
        log(f"Telegram erro: {e}")
        return False


def carregar_historico():
    if os.path.exists(ARQUIVO_HISTORICO):
        with open(ARQUIVO_HISTORICO, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def salvar_historico(dados):
    with open(ARQUIVO_HISTORICO, "w", encoding="utf-8") as f:
        json.dump(dados, f, ensure_ascii=False, indent=2)


def extrair_produtos_da_pagina(html):
    soup = BeautifulSoup(html, "html.parser")
    produtos = []
    cards = soup.select("div[data-asin]:not([data-asin=''])")

    for card in cards:
        try:
            asin = card.get("data-asin", "").strip()
            if not asin or len(asin) < 5:
                continue

            nome_el = card.select_one("h2 span, h2 a span")
            if not nome_el:
                continue
            nome = nome_el.get_text().strip()
            if not nome:
                continue

            preco_el = card.select_one(".a-price .a-offscreen")
            if not preco_el:
                continue
            preco_str = (preco_el.get_text()
                .replace("R$", "").replace("\xa0", "")
                .replace(" ", "").replace(".", "").replace(",", ".").strip()
            )
            try:
                preco = float(preco_str)
                if preco < 10 or preco > 5000:
                    continue
            except ValueError:
                continue

            preco_original_el = card.select_one(".a-text-price .a-offscreen")
            preco_original = None
            if preco_original_el:
                try:
                    orig_str = (preco_original_el.get_text()
                        .replace("R$", "").replace("\xa0", "")
                        .replace(" ", "").replace(".", "").replace(",", ".").strip()
                    )
                    preco_original = float(orig_str)
                except ValueError:
                    pass

            badge_el = card.select_one(".a-badge-text")
            badge = badge_el.get_text().strip() if badge_el else None

            produtos.append({
                "asin": asin,
                "nome": nome[:80],
                "preco": preco,
                "preco_original": preco_original,
                "badge": badge,
                "url": f"https://www.amazon.com.br/dp/{asin}",
            })

        except Exception:
            continue

    return produtos


def varrer_categoria(categoria):
    global bloqueios_consecutivos
    todos_produtos = []
    url_base = categoria["url"]
    log(f"  Varrendo: {categoria['nome']}")

    for pagina in range(1, PAGINAS_VARRER + 1):
        url = f"{url_base}&page={pagina}"
        tentativas = 0

        while tentativas < 3:
            try:
                pausa_humana(4, 10)
                r = requests.get(url, headers=headers_aleatorios(), timeout=20)

                # Bloqueio detectado
                if r.status_code in [503, 429] or "captcha" in r.text.lower() or "robot" in r.text.lower():
                    bloqueios_consecutivos += 1
                    espera = min(60 * (2 ** tentativas), 600)  # backoff exponencial: 60s, 120s, até 600s
                    log(f"  ⚠️  Bloqueio detectado (pág {pagina}, tentativa {tentativas+1}). Aguardando {espera}s...")

                    if bloqueios_consecutivos >= MAX_BLOQUEIOS:
                        enviar_telegram(
                            f"⚠️ <b>VinilOFF Bot — Atenção!</b>\n\n"
                            f"A Amazon bloqueou {bloqueios_consecutivos}x seguidas.\n"
                            f"O bot vai pausar por 15 minutos e tentar novamente.\n\n"
                            f"<i>{datetime.now().strftime('%d/%m %H:%M')}</i>"
                        )
                        bloqueios_consecutivos = 0
                        time.sleep(900)  # pausa longa de 15 min
                    else:
                        time.sleep(espera)

                    tentativas += 1
                    continue

                # Sucesso
                bloqueios_consecutivos = 0
                produtos = extrair_produtos_da_pagina(r.text)

                if not produtos:
                    log(f"  Página {pagina}: sem produtos. Parando categoria.")
                    return todos_produtos

                todos_produtos.extend(produtos)
                log(f"  Página {pagina}: {len(produtos)} produtos")
                break

            except requests.exceptions.Timeout:
                log(f"  Timeout na página {pagina}. Tentando novamente...")
                tentativas += 1
                time.sleep(10)

            except Exception as e:
                log(f"  Erro: {e}")
                tentativas += 1
                time.sleep(10)

        if tentativas >= 3:
            log(f"  Página {pagina} falhou após 3 tentativas. Pulando.")

    return todos_produtos


def analisar_e_alertar(produtos, historico):
    alertas_promocao = []
    alertas_queda = []

    for p in produtos:
        asin = p["asin"]
        preco_atual = p["preco"]
        hist = historico.get(asin)

        if hist is None:
            historico[asin] = {
                "nome": p["nome"],
                "preco_maximo": preco_atual,
                "preco_minimo": preco_atual,
                "preco_anterior": preco_atual,
                "preco_atual": preco_atual,
                "url": p["url"],
                "primeira_vez": datetime.now().strftime("%d/%m/%Y %H:%M"),
                "atualizado": datetime.now().strftime("%d/%m/%Y %H:%M"),
            }
            continue

        preco_anterior = hist.get("preco_atual", preco_atual)
        preco_maximo   = hist.get("preco_maximo", preco_atual)

        historico[asin].update({
            "nome": p["nome"],
            "preco_anterior": preco_anterior,
            "preco_atual": preco_atual,
            "preco_minimo": min(hist.get("preco_minimo", preco_atual), preco_atual),
            "preco_maximo": max(preco_maximo, preco_atual),
            "url": p["url"],
            "atualizado": datetime.now().strftime("%d/%m/%Y %H:%M"),
        })

        if preco_atual >= preco_anterior:
            continue

        desconto_pct = round(((preco_maximo - preco_atual) / preco_maximo) * 100) if preco_maximo > 0 else 0
        queda_atual  = round(((preco_anterior - preco_atual) / preco_anterior) * 100)

        if desconto_pct >= DESCONTO_MINIMO:
            alertas_promocao.append({
                **p,
                "preco_anterior": preco_anterior,
                "preco_maximo": preco_maximo,
                "desconto_pct": desconto_pct,
                "queda_atual": queda_atual,
            })
        elif queda_atual >= 5:
            alertas_queda.append({
                **p,
                "preco_anterior": preco_anterior,
                "queda_atual": queda_atual,
            })

    return alertas_promocao, alertas_queda


def enviar_alertas(alertas_promocao, alertas_queda):
    for a in alertas_promocao[:10]:
        msg = (
            f"🔥 <b>PROMOÇÃO DETECTADA!</b>\n\n"
            f"🎵 <b>{a['nome']}</b>\n\n"
            f"📈 Máximo histórico: <s>R$ {a['preco_maximo']:.2f}</s>\n"
            f"💸 Antes: R$ {a['preco_anterior']:.2f}\n"
            f"✅ Agora: <b>R$ {a['preco_atual']:.2f}</b>\n"
            f"📉 <b>{a['desconto_pct']}% abaixo do máximo!</b>\n"
            f"{f'🏷️ {a[chr(98)+chr(97)+chr(100)+chr(103)+chr(101)]}' if a.get('badge') else ''}\n\n"
            f"🛒 <a href=\"{a['url']}\">Ver na Amazon →</a>\n\n"
            f"<i>VinilOFF · {datetime.now().strftime('%d/%m %H:%M')}</i>"
        )
        enviar_telegram(msg)
        time.sleep(1)

    if alertas_queda:
        linhas = "\n".join([
            f"📀 {q['nome'][:40]}\n   R$ {q['preco_anterior']:.2f} → R$ {q['preco_atual']:.2f} (↓{q['queda_atual']}%)"
            for q in alertas_queda[:8]
        ])
        enviar_telegram(
            f"📉 <b>Quedas detectadas:</b>\n\n{linhas}\n\n"
            f"<i>Abaixo do critério de {DESCONTO_MINIMO}%, mas vale conferir!</i>"
        )


def varredura_completa():
    log("━" * 50)
    log("🔍 Iniciando varredura...")

    historico = carregar_historico()
    todos_produtos = []

    for categoria in CATEGORIAS:
        produtos = varrer_categoria(categoria)
        todos_produtos.extend(produtos)
        time.sleep(random.uniform(8, 15))  # pausa entre categorias

    # Remove duplicatas
    vistos = set()
    unicos = [p for p in todos_produtos if p["asin"] not in vistos and not vistos.add(p["asin"])]

    log(f"📀 {len(unicos)} produtos únicos encontrados")

    alertas_promocao, alertas_queda = analisar_e_alertar(unicos, historico)
    salvar_historico(historico)

    log(f"🔥 Promoções: {len(alertas_promocao)} | 📉 Quedas: {len(alertas_queda)}")

    if alertas_promocao or alertas_queda:
        enviar_alertas(alertas_promocao, alertas_queda)
    else:
        log("Nenhuma promoção nova.")

    log(f"Próxima varredura em {INTERVALO_MINUTOS} min.")
    log("━" * 50 + "\n")


def gerar_planilha():
    """Gera planilha Excel com todos os produtos monitorados"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        if not os.path.exists(ARQUIVO_HISTORICO):
            return None

        with open(ARQUIVO_HISTORICO, "r", encoding="utf-8") as f:
            historico = json.load(f)

        COR_HEADER  = "2C1A0E"
        COR_TEXTO   = "E8DDD0"
        COR_ACCENT  = "C8541A"
        COR_PROMO   = "FFF3EC"
        COR_FORTE   = "FFE0CC"
        COR_PAR     = "F5F0EB"

        def borda():
            s = Side(style="thin", color="DDCCBB")
            return Border(left=s, right=s, top=s, bottom=s)

        # Ordena por desconto
        itens = []
        for asin, d in historico.items():
            preco_atual  = d.get("preco_atual", 0)
            preco_maximo = d.get("preco_maximo", preco_atual)
            preco_minimo = d.get("preco_minimo", preco_atual)
            desconto     = round(((preco_maximo - preco_atual) / preco_maximo) * 100) if preco_maximo > 0 else 0
            itens.append({
                "nome": d.get("nome", ""), "preco_atual": preco_atual,
                "preco_maximo": preco_maximo, "preco_minimo": preco_minimo,
                "desconto": desconto, "url": d.get("url", ""),
            })
        itens.sort(key=lambda x: x["desconto"], reverse=True)

        wb = Workbook()

        # ── ABA 1: TODOS ──
        ws = wb.active
        ws.title = "Todos os Produtos"

        ws.merge_cells("A1:H1")
        ws["A1"] = "VinilOFF — Produtos Monitorados"
        ws["A1"].font = Font(name="Arial", bold=True, size=14, color=COR_TEXTO)
        ws["A1"].fill = PatternFill("solid", fgColor=COR_HEADER)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        ws.merge_cells("A2:H2")
        ws["A2"] = f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')} · {len(itens)} produtos"
        ws["A2"].font = Font(name="Arial", size=10, color="7A5C4A")
        ws["A2"].fill = PatternFill("solid", fgColor="3D2510")
        ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 18

        headers = ["#", "Produto", "Preço Atual", "Preço Máximo", "Preço Mínimo", "Desconto %", "Status", "Link"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=3, column=col, value=h)
            c.font = Font(name="Arial", bold=True, size=10, color=COR_TEXTO)
            c.fill = PatternFill("solid", fgColor=COR_ACCENT)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = borda()
        ws.row_dimensions[3].height = 22

        for i, item in enumerate(itens):
            row = i + 4
            cor = COR_FORTE if item["desconto"] >= 20 else COR_PROMO if item["desconto"] >= 10 else COR_PAR if i % 2 == 0 else "FFFFFF"
            fill = PatternFill("solid", fgColor=cor)

            dados = [
                (i+1, "center", False, "7A5C4A"),
                (item["nome"], "left", item["desconto"] >= 10, "000000"),
                (item["preco_atual"], "center", True, COR_ACCENT if item["desconto"] >= 10 else "000000"),
                (item["preco_maximo"], "center", False, "999999"),
                (item["preco_minimo"], "center", False, "1A3A2A"),
                (f"=(D{row}-C{row})/D{row}", "center", item["desconto"] >= 10, COR_ACCENT if item["desconto"] >= 10 else "666666"),
                ("🔥 PROMOÇÃO" if item["desconto"] >= 20 else "📉 Em queda" if item["desconto"] >= 10 else "💤 Normal", "center", False, "000000"),
                (item["url"], "center", False, "1155CC"),
            ]

            for col, (val, align, bold, color) in enumerate(dados, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.font = Font(name="Arial", size=10, bold=bold, color=color)
                c.fill = fill
                c.border = borda()
                c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=(col==2))
                if col in [3, 4, 5]:
                    c.number_format = 'R$ #,##0.00'
                if col == 6:
                    c.number_format = '0%'
                if col == 8:
                    c.hyperlink = val
                    c.font = Font(name="Arial", size=9, color="1155CC", underline="single")
            ws.row_dimensions[row].height = 20

        for col, width in zip("ABCDEFGH", [5, 42, 14, 14, 14, 12, 16, 35]):
            ws.column_dimensions[col].width = width
        ws.freeze_panes = "A4"

        # ── ABA 2: PROMOÇÕES ──
        ws2 = wb.create_sheet("🔥 Promoções Ativas")
        ws2.merge_cells("A1:G1")
        ws2["A1"] = f"VinilOFF — Promoções Ativas · {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws2["A1"].font = Font(name="Arial", bold=True, size=13, color=COR_TEXTO)
        ws2["A1"].fill = PatternFill("solid", fgColor=COR_ACCENT)
        ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[1].height = 28

        h2 = ["#", "Produto", "Preço Atual", "Preço Máximo", "Desconto", "Economia", "Link"]
        for col, h in enumerate(h2, 1):
            c = ws2.cell(row=2, column=col, value=h)
            c.font = Font(name="Arial", bold=True, size=10, color=COR_TEXTO)
            c.fill = PatternFill("solid", fgColor=COR_HEADER)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = borda()
        ws2.row_dimensions[2].height = 22

        promos = [x for x in itens if x["desconto"] >= 10]
        for i, item in enumerate(promos):
            row = i + 3
            fill = PatternFill("solid", fgColor=COR_FORTE if item["desconto"] >= 20 else COR_PROMO)
            vals = [i+1, item["nome"], item["preco_atual"], item["preco_maximo"],
                    f"=(D{row}-C{row})/D{row}", f"=D{row}-C{row}", item["url"]]
            for col, val in enumerate(vals, 1):
                c = ws2.cell(row=row, column=col, value=val)
                c.fill = fill
                c.border = borda()
                c.alignment = Alignment(horizontal="center" if col != 2 else "left", vertical="center", wrap_text=(col==2))
                bold = col in [2, 3, 5, 6]
                color = COR_ACCENT if col in [3, 5] else "1A3A2A" if col == 6 else "1155CC" if col == 7 else "000000"
                c.font = Font(name="Arial", size=10, bold=bold, color=color)
                if col in [3, 4, 6]:
                    c.number_format = 'R$ #,##0.00'
                if col == 5:
                    c.number_format = '0%'
                if col == 7:
                    c.hyperlink = val
                    c.font = Font(name="Arial", size=9, color="1155CC", underline="single")
            ws2.row_dimensions[row].height = 20

        for col, width in zip("ABCDEFG", [5, 42, 14, 14, 12, 14, 35]):
            ws2.column_dimensions[col].width = width
        ws2.freeze_panes = "A3"

        caminho = "viniloff_produtos.xlsx"
        wb.save(caminho)
        log(f"📊 Planilha gerada: {len(itens)} produtos, {len(promos)} em promoção")
        return caminho

    except Exception as e:
        log(f"Erro ao gerar planilha: {e}")
        return None


def enviar_documento_telegram(caminho_arquivo, caption=""):
    """Envia arquivo para o Telegram"""
    try:
        url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendDocument"
        with open(caminho_arquivo, "rb") as f:
            r = requests.post(url, data={
                "chat_id": TELEGRAM_CHAT_ID,
                "caption": caption,
                "parse_mode": "HTML",
            }, files={"document": f}, timeout=30)
        return r.status_code == 200
    except Exception as e:
        log(f"Erro ao enviar documento: {e}")
        return False


def verificar_comandos():
    """Fica escutando comandos do Telegram em segundo plano"""
    ultimo_update = 0
    log("📱 Escutando comandos do Telegram...")

    while True:
        try:
            url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/getUpdates"
            r = requests.get(url, params={"offset": ultimo_update + 1, "timeout": 30}, timeout=35)

            if r.status_code != 200:
                time.sleep(10)
                continue

            data = r.json()
            for update in data.get("result", []):
                ultimo_update = update["update_id"]
                msg = update.get("message", {})
                texto = msg.get("text", "").strip().lower()
                chat_id = str(msg.get("chat", {}).get("id", ""))

                # Só responde ao dono do bot
                if chat_id != str(TELEGRAM_CHAT_ID):
                    continue

                if texto == "/planilha":
                    log("📊 Comando /planilha recebido!")
                    enviar_telegram("⏳ Gerando planilha, aguarde...")
                    caminho = gerar_planilha()
                    if caminho and os.path.exists(caminho):
                        historico = carregar_historico()
                        total = len(historico)
                        promos = len([x for x in historico.values()
                                      if x.get("preco_maximo", 0) > 0 and
                                      round(((x.get("preco_maximo", 0) - x.get("preco_atual", 0)) / x.get("preco_maximo", 1)) * 100) >= 10])
                        caption = (
                            f"📊 <b>VinilOFF — Planilha atualizada</b>\n\n"
                            f"📀 {total} produtos monitorados\n"
                            f"🔥 {promos} em promoção\n"
                            f"<i>{datetime.now().strftime('%d/%m/%Y %H:%M')}</i>"
                        )
                        if enviar_documento_telegram(caminho, caption):
                            log("✅ Planilha enviada no Telegram!")
                        else:
                            enviar_telegram("❌ Erro ao enviar planilha.")
                    else:
                        enviar_telegram("❌ Nenhum dado ainda. Aguarde a primeira varredura.")

                elif texto == "/status":
                    historico = carregar_historico()
                    total = len(historico)
                    promos = len([x for x in historico.values()
                                  if x.get("preco_maximo", 0) > 0 and
                                  round(((x.get("preco_maximo", 0) - x.get("preco_atual", 0)) / x.get("preco_maximo", 1)) * 100) >= 10])
                    enviar_telegram(
                        f"📊 <b>Status do VinilOFF Bot</b>\n\n"
                        f"✅ Online e funcionando\n"
                        f"📀 {total} produtos monitorados\n"
                        f"🔥 {promos} em promoção agora\n"
                        f"⏱️ Varredura a cada {INTERVALO_MINUTOS} min\n"
                        f"🎯 Alerta com desconto ≥ {DESCONTO_MINIMO}%\n\n"
                        f"<i>{datetime.now().strftime('%d/%m/%Y %H:%M')}</i>"
                    )

                elif texto == "/ajuda":
                    enviar_telegram(
                        f"🤖 <b>Comandos do VinilOFF Bot</b>\n\n"
                        f"/planilha — Baixar planilha Excel com todos os produtos\n"
                        f"/status — Ver quantos produtos estão em promoção\n"
                        f"/ajuda — Ver esta mensagem\n\n"
                        f"<i>Alertas automáticos chegam quando um vinil entra em promoção!</i>"
                    )

        except Exception as e:
            log(f"Erro no listener: {e}")
            time.sleep(15)


def main():
    log("🎵 VinilOFF Bot iniciado!")
    log(f"⏱️  Intervalo: {INTERVALO_MINUTOS} min | 🎯 Desconto mínimo: {DESCONTO_MINIMO}%")

    enviar_telegram(
        f"🎵 <b>VinilOFF Bot atualizado!</b>\n\n"
        f"🛡️ Versão anti-bloqueio ativa\n"
        f"🔍 {len(CATEGORIAS)} categorias · {PAGINAS_VARRER} páginas cada\n"
        f"🎯 Alerta com desconto ≥ <b>{DESCONTO_MINIMO}%</b>\n"
        f"⏱️  A cada <b>{INTERVALO_MINUTOS} minutos</b>\n\n"
        f"Comandos disponíveis:\n"
        f"/planilha — baixar planilha Excel\n"
        f"/status — ver resumo\n"
        f"/ajuda — ver todos os comandos\n\n"
        f"Monitoramento ativo! 🔥"
    )

    # Inicia listener de comandos em segundo plano
    thread = threading.Thread(target=verificar_comandos, daemon=True)
    thread.start()
    log("📱 Listener de comandos iniciado!")

    while True:
        try:
            varredura_completa()
            time.sleep(INTERVALO_MINUTOS * 60)
        except KeyboardInterrupt:
            log("Bot encerrado.")
            enviar_telegram("⏹️ <b>VinilOFF Bot encerrado.</b>")
            break
        except Exception as e:
            log(f"Erro inesperado: {e}. Tentando em 10 min...")
            time.sleep(600)


if __name__ == "__main__":
    main()
